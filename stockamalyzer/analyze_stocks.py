import requests
import yfinance as yf
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import os
import numpy as np
from datetime import datetime
from openpyxl.drawing.image import Image as XLImage
from matplotlib.gridspec import GridSpec
from matplotlib.patches import Rectangle
from mplfinance.original_flavor import candlestick_ohlc
import matplotlib.dates as mdates
import talib
from fpdf import FPDF
from sklearn.linear_model import LinearRegression
import re

# === CONFIG ===
VAL_THRESHOLD = 0.02      # ±2% threshold
SHIFT_6M = 126            # ~6 months ≈ 126 trading days
SHIFT_12M = 252           # ~12 months ≈ 252 trading days
SHIFT_24M = 504           # ~24 months ≈ 504 trading days

tplt = plt.get_cmap('tab10')
PALETTE = tplt.colors
ACCENT = '#ff6f00'
plt.style.use('default')

def safe_text(s):
    return s.replace('—', '-').replace('–', '-')

def parse_period(p: str):
    m = re.match(r'^(\d+)([ywdmo]+)$', p.lower())
    if not m:
        return 1260, 'd'
    n, u = int(m.group(1)), m.group(2)
    if u == 'y': return n * 252, 'd'
    if u == 'mo': return n, 'mo'
    if u == 'w': return n, 'wk'
    if u == 'd': return n, 'd'
    return 1260, 'd'

def get_return_periods(days: int):
    if days <= SHIFT_24M:
        return SHIFT_6M, SHIFT_12M, '6M Return %', '12M Return %'
    else:
        return SHIFT_12M, SHIFT_24M, '1Y Return %', '2Y Return %'

def draw_pattern_visual(ax, df, idx, pattern):
    bar = 0.6
    involved = ([idx-2, idx-1, idx] if '3' in pattern or 'MORNINGSTAR' in pattern else [idx-1, idx])
    for i in involved:
        if 0 <= i < len(df):
            x0 = df['Date_Num'].iat[i] - bar/2
            y0 = df['Low'].iat[i]
            h = df['High'].iat[i] - y0
            ax.add_patch(Rectangle((x0, y0), bar, h, edgecolor=ACCENT, facecolor=ACCENT, alpha=0.4))

def detect_valid_channels(df, ax, lookback=50, stride=5, min_slope=0.005):
    if len(df) < lookback:
        return
    last_end = -1
    up_flag = down_flag = False
    for start in range(0, len(df)-lookback, stride):
        if start < last_end:
            continue
        end = start + lookback
        x = np.arange(lookback).reshape(-1,1)
        y_h = df['High'].iloc[start:end].values.reshape(-1,1)
        y_l = df['Low'].iloc[start:end].values.reshape(-1,1)
        close = df['Close'].iloc[start:end].values
        rh = LinearRegression().fit(x, y_h)
        rl = LinearRegression().fit(x, y_l)
        sh, sl = rh.coef_[0][0], rl.coef_[0][0]
        upper = rh.predict(x).flatten()
        lower = rl.predict(x).flatten()
        width_var = np.std(upper - lower)
        within = np.mean((close >= lower) & (close <= upper))
        is_up = sh > min_slope and sl > min_slope
        is_down = sh < -min_slope and sl < -min_slope
        if ((is_up and not up_flag) or (is_down and not down_flag)) and width_var<4 and within>0.4:
            label = 'Upward Channel' if is_up else 'Downward Channel'
            color, hatch = PALETTE[3], '///'
            dates = df['Date_Num'].iloc[start:end]
            ax.plot(dates, upper, '--', lw=1.2, color=color, label=label)
            ax.plot(dates, lower, '--', lw=1.2, color=color)
            ax.fill_between(dates, lower, upper, color=color, alpha=0.2, hatch=hatch)
            last_end = end
            if is_up: up_flag = True
            else: down_flag = True

def analyze_stock(ticker: str, period: str, freq_str: str):
    freq_map = {'daily': '1d', 'weekly': '1wk', 'monthly': '1mo'}
    interval = freq_map.get(freq_str.lower(), '1d')
    df = yf.Ticker(ticker).history(
        period='max', interval=interval,
        auto_adjust=False, prepost=True
    ).dropna(subset=['Open','High','Low','Close','Volume'])
    df['Date_Num'] = mdates.date2num(df.index.to_pydatetime())
    df['MA20'] = df['Close'].rolling(20).mean()
    df['MA50'] = df['Close'].rolling(50).mean()
    df['MA100'] = df['Close'].rolling(100).mean()
    d = df['Close'].diff(); g = d.clip(lower=0); l = -d.clip(upper=0)
    df['RSI'] = 100 - 100/(1 + g.rolling(14).mean()/l.rolling(14).mean())
    e12 = df['Close'].ewm(span=12).mean(); e26 = df['Close'].ewm(span=26).mean()
    df['MACD'] = e12 - e26
    df['Signal'] = df['MACD'].ewm(span=9).mean()
    pc, pu = parse_period(period)
    days = pc if pu=='d' else pc*21 if pu=='mo' else pc*5
    days = min(days, len(df))
    dfp = df.tail(days)
    matches = []
    for name in dir(talib):
        if not name.startswith('CDL'): continue
        res = getattr(talib, name)(dfp['Open'], dfp['High'], dfp['Low'], dfp['Close'])
        idxs = np.where(res != 0)[0]
        for i in idxs:
            matches.append({'name': name.replace('CDL',''), 'index': i, 'date': dfp.index[i], 'strength': int(res.iloc[i])})
    if matches:
        best = max(matches, key=lambda x:(x['date'], abs(x['strength'])))
        detected, strength = best['name'], f"{best['name']}={best['strength']}"
    else:
        detected, strength = 'None',''
    last_rsi = df['RSI'].iloc[-1]
    if detected.upper() in ['HAMMER','ENGULFING','MORNINGSTAR'] or last_rsi < 35:
        classification = 'Bullish'
    elif detected.upper() in ['SHOOTINGSTAR','HANGINGMAN'] or last_rsi > 65:
        classification = 'Bearish'
    else:
        classification = 'Neutral'
    os.makedirs('charts', exist_ok=True)
    fname = f"charts/{ticker}_{datetime.now():%Y%m%d%H%M%S}.png"
    fig = plt.figure(figsize=(10,6))
    gs = GridSpec(4,1, height_ratios=[3,1,1,1])
    ax1 = fig.add_subplot(gs[0])
    candlestick_ohlc(ax1, dfp[['Date_Num','Open','High','Low','Close']].values,
                     width=0.6, colorup=PALETTE[0], colordown=PALETTE[1])
    ax1.plot(dfp['Date_Num'], dfp['MA20'], label='MA20')
    ax1.plot(dfp['Date_Num'], dfp['MA50'], linestyle='--', label='MA50')
    ax1.plot(dfp['Date_Num'], dfp['MA100'], linestyle=':', label='MA100')
    ax1.set_title(f"{ticker} | {freq_str.capitalize()} ({period})")
    ax1.legend(fontsize=8)
    detect_valid_channels(dfp, ax1)
    if detected != 'None':
        i = best['index']
        ax1.annotate(detected,
                     xy=(dfp['Date_Num'].iat[i], dfp['High'].iat[i]),
                     xytext=(dfp['Date_Num'].iat[i], dfp['High'].iat[i]*1.05),
                     bbox={'boxstyle':'round','fc':'white','ec':'black'},
                     arrowprops={'arrowstyle':'->'})
        draw_pattern_visual(ax1, dfp, i, detected)
    ax1.grid(True, linestyle=':', alpha=0.4)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    ax2 = fig.add_subplot(gs[1], sharex=ax1)
    ax2.bar(dfp['Date_Num'], dfp['Volume'], color='gray')
    ax2.legend(['Volume'], fontsize=8)
    ax3 = fig.add_subplot(gs[2], sharex=ax1)
    ax3.plot(dfp['Date_Num'], dfp['RSI'], label='RSI'); ax3.axhline(70,ls='--'); ax3.axhline(30,ls='--'); ax3.legend(fontsize=8)
    ax4 = fig.add_subplot(gs[3], sharex=ax1)
    ax4.plot(dfp['Date_Num'], dfp['MACD'], label='MACD'); ax4.plot(dfp['Date_Num'], dfp['Signal'], label='Signal'); ax4.legend(fontsize=8)
    fig.autofmt_xdate(); plt.tight_layout(); fig.savefig(fname, dpi=80); plt.close(fig)
    return classification, fname, df['RSI'].iloc[-1], df['MACD'].iloc[-1], df['Signal'].iloc[-1], detected, strength, df

# === EXCEL & PDF SETUP ===
wb = openpyxl.load_workbook('stocks.xlsx')
ws_current = wb['Current']
ws_history = wb['History'] if 'History' in wb.sheetnames else wb.create_sheet('History')
if ws_history.max_row == 1:
    ws_history.append(['Date','Ticker','Period','Freq','Classification','RSI','MACD','Signal','Pattern'])

today = datetime.now().strftime('%Y-%m-%d')
pdf = FPDF()
pdf.set_auto_page_break(True, margin=15)
pdf.add_page()
pdf.set_font('Arial','B',16)
pdf.cell(0,10,'Stock Pattern Summary Report',0,1,'C')
pdf.ln(5)
pdf.set_font('Arial','',12)
pdf.cell(0,6,f"Date: {today}",0,1)

summary_rows = []

# === MAIN LOOP ===
for row in ws_current.iter_rows(min_row=2, max_col=3):
    tkr = row[0].value
    per = row[1].value or '5y'
    freq = row[2].value or 'daily'
    if not tkr:
        continue
    tkr = tkr.upper()

    try:
        cl, chart_path, rsi, macd, sig, patt, strg, df_full = analyze_stock(tkr, per, freq)

        ws_current.cell(row=row[0].row, column=4).value = cl
        ws_current.cell(row=row[0].row, column=5).value = chart_path
        ws_current.cell(row=row[0].row, column=6).value = patt
        ws_current.cell(row=row[0].row, column=7).value = strg

        if os.path.exists(chart_path):
            img = XLImage(chart_path)
            img.width, img.height = 180, 120
            ws_current.add_image(img, f'H{row[0].row}')

        ws_history.append([today, tkr, per, freq, cl, round(rsi, 2), round(macd, 2), round(sig, 2), patt])
        summary_rows.append((tkr, per, freq, cl, patt, chart_path, df_full))

    except Exception as e:
        ws_current.cell(row=row[0].row, column=4).value = f"Error: {e}"

# === OVERVIEW SUMMARY TABLE ===
pdf.add_page()
pdf.set_font('Arial', 'B', 14)
pdf.cell(0, 10, 'Stock Overview Summary', 0, 1, 'C')
pdf.ln(4)
pdf.set_font('Arial', 'B', 11)
for hdr in ['Ticker', 'Period', 'Freq', 'Classification', 'Pattern']:
    pdf.cell(35, 6, hdr, 1, 0, 'C')
pdf.ln()
pdf.set_font('Arial', '', 10)
for tkr, per, freq, cl, patt, _, _ in summary_rows:
    for v in [tkr, per, freq, cl, patt]:
        pdf.cell(35, 6, str(v), 1, 0, 'C')
    pdf.ln()

# === DETAIL PAGES ===
for tkr, per, freq, cl, patt, chart_path, df_full in summary_rows:
    if patt == 'None':
        continue

    pdf.add_page()
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f"{tkr} | Chart & Indicators", 0, 1, 'C')
    if os.path.exists(chart_path):
        pdf.image(chart_path, 10, 30, w=180)

    # === VBA EV/EBITDA Chart ===
    ev_chart_path = f"C:/Users/akats/OneDrive/Desktop/stockamalyzer/outputs/{tkr}_NTM_EV_EBITDA_Chart.png"
    pdf.add_page()
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f"{tkr} EV/EBITDA (VBA Export)", 0, 1, 'C')
    if os.path.exists(ev_chart_path):
        pdf.image(ev_chart_path, 10, 30, w=180)
    else:
        pdf.cell(0, 10, "EV/EBITDA chart not found.", 0, 1, 'C')

    # === Pattern Occurrences ===
    pc, pu = parse_period(per)
    days_back = pc if pu == 'd' else pc * 21 if pu == 'mo' else pc * 5
    df_period = df_full.tail(days_back)
    if df_period.index.tz is not None:
        df_period.index = df_period.index.tz_localize(None)

    talib_func = getattr(talib, f"CDL{patt}", None)
    occ_idx = []
    if talib_func:
        try:
            res2 = talib_func(
                df_period['Open'],
                df_period['High'],
                df_period['Low'],
                df_period['Close']
            )
            occ_idx = np.where(res2 != 0)[0]
        except Exception as e:
            print(f"[WARN] Pattern Occurrence error for {tkr}: {e}")

    fig2, ax2 = plt.subplots(figsize=(8, 4))
    ax2.plot(df_period.index, df_period['Close'], lw=1.5)
    if len(occ_idx) > 0:
        ax2.scatter(df_period.index[occ_idx], df_period['Close'].iloc[occ_idx], color=ACCENT, s=50)
    ax2.set_title(f"{tkr} Occurrences of {patt}", fontsize=12)
    ax2.grid(True, linestyle=':', alpha=0.5)
    fig2.autofmt_xdate()
    occ_path = f"charts/{tkr}_occ_{datetime.now():%Y%m%d%H%M%S}.png"
    fig2.savefig(occ_path, dpi=80)
    plt.close(fig2)

    pdf.add_page()
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f"{tkr} Pattern Occurrences: {len(occ_idx)}", 0, 1, 'C')
    pdf.image(occ_path, 10, 30, w=180)
    os.remove(occ_path)

    # === Price Matches Table ===
    short_p, long_p, sp_lbl, lp_lbl = get_return_periods(days_back)
    curr_price = df_period['Close'].iloc[-1]

    msk = (df_period['Close'] - curr_price).abs() / curr_price <= VAL_THRESHOLD
    vals = []

    for d in df_period.index[msk]:
        pos = df_period.index.get_loc(d)
        if pos + short_p >= len(df_period) or pos + long_p >= len(df_period):
            continue

        p0 = df_period['Close'].iloc[pos]
        ps = df_period['Close'].shift(-short_p).iloc[pos]
        pl = df_period['Close'].shift(-long_p).iloc[pos]

        if np.isnan(ps) or np.isnan(pl):
            continue

        vals.append(((ps - p0) / p0 * 100, (pl - p0) / p0 * 100))

    avg_sp = np.nanmean([v[0] for v in vals]) if vals else 0.0
    avg_lp = np.nanmean([v[1] for v in vals]) if vals else 0.0

    pdf.add_page()
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f"{tkr} Price Matches (±{VAL_THRESHOLD*100:.0f}% of {curr_price:.2f})", 0, 1, 'C')
    pdf.ln(4)

    pdf.set_font('Arial', 'B', 10)
    pdf.cell(50, 5, 'Price', 1, 0, 'C')
    pdf.cell(50, 5, sp_lbl, 1, 0, 'C')
    pdf.cell(50, 5, lp_lbl, 1, 0, 'C')
    pdf.ln()

    pdf.set_font('Arial', '', 10)
    pdf.cell(50, 5, f"{curr_price:.2f}", 1, 0, 'C')
    pdf.cell(50, 5, f"{avg_sp:.2f}", 1, 0, 'C')
    pdf.cell(50, 5, f"{avg_lp:.2f}", 1, 0, 'C')
    pdf.ln()

    # === Top 20 Patterns & Returns ===
    pattern_counts = {}
    pattern_returns = {}

    for name in dir(talib):
        if not name.startswith('CDL'):
            continue
        res = getattr(talib, name)(df_period['Open'], df_period['High'], df_period['Low'], df_period['Close'])
        idxs = np.where(res != 0)[0]
        f_s, f_l = [], []
        for ix in idxs:
            if ix + short_p >= len(df_period) or ix + long_p >= len(df_period):
                continue
            p0 = df_period['Close'].iloc[ix]
            ps = df_period['Close'].shift(-short_p).iloc[ix]
            pl = df_period['Close'].shift(-long_p).iloc[ix]
            if np.isnan(ps) or np.isnan(pl):
                continue
            f_s.append((ps - p0) / p0 * 100)
            f_l.append((pl - p0) / p0 * 100)
        if idxs.size > 0:
            p = name.replace('CDL', '')
            pattern_counts[p] = len(idxs)
            pattern_returns[p] = {
                'short': np.nanmean(f_s) if f_s else 0.0,
                'long': np.nanmean(f_l) if f_l else 0.0
            }

    top20 = sorted(pattern_counts.items(), key=lambda x: x[1], reverse=True)[:20]
    pdf.add_page()
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, f"{tkr} Top 20 Patterns & Returns", 0, 1, 'C')

    pdf.set_font('Arial', 'B', 10)
    for col, w in zip(['Pattern', 'Count', sp_lbl, lp_lbl], [60, 25, 30, 30]):
        pdf.cell(w, 5, col, 1, 0, 'C')
    pdf.ln()

    pdf.set_font('Arial', '', 10)
    for p, count in top20:
        sh = pattern_returns[p]['short']
        ln = pattern_returns[p]['long']
        pdf.cell(60, 5, p, 1, 0, 'C')
        pdf.cell(25, 5, str(count), 1, 0, 'C')
        pdf.cell(30, 5, f"{sh:.2f}", 1, 0, 'C')
        pdf.cell(30, 5, f"{ln:.2f}", 1, 0, 'C')
        pdf.ln()

from PIL import Image  # at the top of your script

image_path = f"C:/Users/akats/OneDrive/Desktop/stockamalyzer/outputs/{tkr}_Forward_Returns_Table.png"

pdf.add_page()
pdf.set_font('Arial', 'B', 14)
pdf.cell(0, 10, f"{tkr} Forward Returns Table", ln=True, align='C')
pdf.ln(5)

if os.path.exists(image_path):
    with Image.open(image_path) as img:
        img_w, img_h = img.size
        dpi = 96
        img_w_mm = img_w / dpi * 25.4
        img_h_mm = img_h / dpi * 25.4

    max_w = 180  # A4 width minus margins
    max_h = 250  # A4 height minus margins
    scale_w = max_w / img_w_mm
    scale_h = max_h / img_h_mm
    scale = min(scale_w, scale_h, 1.0)

    pdf_w = img_w_mm * scale
    pdf_h = img_h_mm * scale

    x = (210 - pdf_w) / 2
    pdf.image(image_path, x=x, y=None, w=pdf_w, h=pdf_h)

else:
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f"[X] Could not find: {image_path}", ln=True, align='C')

# === FINALIZE ===
wb.save('stocks.xlsx')
pdf.output('stock_report.pdf', 'F')
print("✅ Stock report generated successfully!")
